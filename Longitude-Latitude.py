import os  # to check and rename files
import pandas
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import numbers


def setup():
    # default name of the file downloaded from https://nwis.waterdata.usgs.gov/usa/nwis/peak.
    if os.path.isfile("./location"):
        os.rename('location', 'location.txt')  # We add the .txt extension to tell the program it is a text file.
        print("Changed file type to text file.")

    if os.path.isfile("./location.txt"):
        print("File Found.")
        return 0
    else:
        print("File not found. Please make sure the file is named 'location.txt' and is in the correct directory.")
        return -1


def removeComments(fileName, n):
    with open(fileName, 'r') as file:
        lines = file.readlines()

    with open(fileName, 'w') as file:
        file.writelines(lines[n:])

    return 0


def main():
    # Setup
    fileFound = setup()
    if fileFound == 0:
        print("Setup Complete.")
    else:
        print("Error 1")
        return 0

    # Find Number of Lines to Remove
    fileName = "location.txt"
    n = 0

    with open(fileName, 'r') as file:
        for currentLine in file:
            if currentLine.startswith("#"):
                n += 1

    # Remove the lines that start with '#'
    if removeComments(fileName, n) == 0:
        print("Comments Removed.")
    else:
        print("Error 2")
        exit()

    print("Ready to convert .txt file to .xlsx")

    # Convert .txt to .xlsx
    df = pandas.read_csv("location.txt", delimiter='\t')
    # pandas.to_numeric(df[6])
    df.to_excel("location.xlsx", index=False)

    workbook = openpyxl.load_workbook("location.xlsx")
    worksheet = workbook["Sheet1"]
    worksheet.delete_rows(2)
    workbook.save("location.xlsx")
    print("Removed Header 2.")

    workbook = load_workbook("location.xlsx")
    worksheet = workbook["Sheet1"]
    row = 2

    print("Reformatting Data Types...")
    # Convert Columns A, B, C, D, E, & F to Text Data Type and G & H to Number Data Type
    while row <= 1400:
        columnA = "A" + str(row)
        columnB = "B" + str(row)
        columnC = "C" + str(row)
        columnD = "D" + str(row)
        columnE = "E" + str(row)
        columnF = "F" + str(row)
        columnG = "G" + str(row)
        columnH = "H" + str(row)
        columnI = "I" + str(row)
        columnJ = "J" + str(row)
        columnK = "K" + str(row)

        worksheet[columnA].number_format = numbers.FORMAT_TEXT
        worksheet[columnB].number_format = numbers.FORMAT_TEXT
        worksheet[columnC].number_format = numbers.FORMAT_TEXT
        worksheet[columnD].number_format = numbers.FORMAT_TEXT
        worksheet[columnE].number_format = numbers.FORMAT_TEXT
        worksheet[columnF].number_format = numbers.FORMAT_TEXT
        worksheet[columnG].number_format = numbers.FORMAT_TEXT
        worksheet[columnH].number_format = numbers.FORMAT_NUMBER
        worksheet[columnI].number_format = numbers.FORMAT_NUMBER
        worksheet[columnJ].number_format = numbers.FORMAT_TEXT
        worksheet[columnK].number_format = numbers.FORMAT_NUMBER

        row += 1
    print("Complete.")

    # Save the workbook
    workbook.save('location.xlsx')
    print("Workbook Saved.")
    return 0


main()
