import os
from openpyxl import load_workbook
import pymannkendall as mk
import statistics
from openpyxl.styles import numbers
import pandas as pd


def MannKendallTest(data, site):
    # print("data", data) # debug

    index = 0

    MKResults = []
    individual = []

    print("Performing Tests...")
    # Import Results into a list (MKResults) of lists (individual)
    # Perform the Mann-Kendall test
    result = mk.original_test(data, alpha=0.05)
    mean = statistics.mean(data)
    median = statistics.median(data)
    stdev = statistics.stdev(data)

    # Extract results
    trend = result.trend
    hypothesis = result.h
    p_value = result.p
    # print("P:", p_value) # debug

    # Add data to individual
    individual.append(site)
    individual.append(trend)
    individual.append(hypothesis)
    individual.append(p_value)
    individual.append(mean)
    individual.append(median)
    individual.append(stdev)

    # Append individual site results to MKResults
    # print("Results: ", individual) # debug
    return individual


def formatResults():
    workbook = load_workbook("../all-sites.xlsx")
    worksheet = workbook["Sheet1"]

    site = []
    AllSites = []
    row = 2

    print("Pulling data for comparison...")
    # Pull Longitude, Latitude, and Site Name for every Site Number (1383 total)
    while row < 1384:
        site_no = "B" + str(row)
        station_nm = "C" + str(row)
        latitude = "D" + str(row)
        longitude = "E" + str(row)

        if worksheet[site_no] is not None:
            site.append(worksheet[site_no].value)
            site.append(worksheet[station_nm].value)
            site.append(worksheet[latitude].value)
            site.append(worksheet[longitude].value)

        AllSites.append(site)
        site = []
        row += 1

    print("Complete.")
    workbook2 = load_workbook("Test-Results.xlsx")
    worksheet2 = workbook2["Sheet1"]
    worksheet2.insert_cols(1)

    # Variable definitions
    row = 1
    columnA = "A" + str(row)
    columnI = "I" + str(row)
    columnJ = "J" + str(row)

    print("Adding titles for new data columns...")
    # Add titles for the new columns
    worksheet2[columnA].value = "Station Name"
    worksheet2[columnI].value = "Latitude"
    worksheet2[columnJ].value = "Longitude"

    print("Complete.")
    workbook2.save('Test-Results.xlsx')
    print("Workbook Saved.")

    # Find matching site numbers
    workingSites = []
    while row < 1400:
        row += 1
        columnB = "B" + str(row)
        # If we're not at the end of the data
        if worksheet2[columnB].value is not None:
            workingSites.append(worksheet2[columnB].value)
    # print(len(workingSites)) # debug

    print("Importing Latitude/Longitude data for dataset...")
    count = 0
    for index in AllSites:
        for i in workingSites:
            # If the site number at index i from working sites is equal to that from all the sites...
            if i == index[0]:
                columnA = "A" + str(count + 2)
                columnI = "I" + str(count + 2)
                columnJ = "J" + str(count + 2)
                # ... add the lat/long data and site name to the correct column in Test-Results.xlsx
                worksheet2[columnA].value = index[1]
                worksheet2[columnI].value = index[2]
                worksheet2[columnJ].value = index[3]
                count += 1
    # print(count) # debug

    print("Complete.")
    workbook2.save('Test-Results.xlsx')
    print("Workbook Saved.")

    return 0


def main():
    # default name of the file downloaded from https://nwis.waterdata.usgs.gov/usa/nwis/peak.
    if os.path.isfile("./dv"):
        os.rename('dv', 'dv.txt')  # We add the .txt extension to tell the program it is a text file.
        print("Changed file type to text file.")

    if os.path.isfile("./dv.txt"):
        print("File Found.")
    else:
        print("File not found. Please make sure the file is named 'peak.txt' and is in the correct directory.")
        return 0

    # Remove all lines in file that begin with an octothorpe
    fileNum = 0
    # Make a list of lines that need to be deleted
    linesToDelete = []
    titleLines = 0
    fileNum += 1

    with open("dv.txt", 'r') as f:
        lines = f.readlines()
        for index, line in enumerate(lines):
            if line.startswith("#"):
                linesToDelete.append(index)
            elif line.startswith("5s"):
                linesToDelete.append(index)

    # print(linesToDelete) # debug

    # Delete the lines in the list linesToDelete
    for index in reversed(linesToDelete):
        del lines[index]

    with open("dv.txt", 'w') as f:
        f.writelines(lines)
        f.write("\n")
        f.write("agency_cd")

    # Definitions
    flood_data = []
    instance = []
    site = 0
    sites = []
    MKResults = []
    dontinclude = 0

    with open("dv.txt", 'r') as file:
        for line in file:
            if line.startswith("agency_cd"):
                # Gather data to send to Mann Kendall Test Function
                for item in instance:
                    # print(item) #debug
                    try:
                        flood_data.append(float(item[7]))
                    except IndexError:
                        print("No Data. Skipping Instance.")
                if len(flood_data) >= 1:
                    sites.append(instance[0][1])
                    # Go to Mann Kendall Test Function
                    Results = MannKendallTest(flood_data, sites[0-1])
                    MKResults.append(Results)
                else:
                    print("Not enough data for this site number to include in report.")
                    dontinclude += 1

                # Reset Lists for Next Station
                instance = []
                # print(flood_data) # debug
                flood_data = []
                continue
            else:
                # Split lines to read the column of data needed
                stripped_line = line.strip()
                items = stripped_line.split('\t')
                instance.append(items)

    # print(MKResults) # debug
    # print(Results) # debug
    print("Importing Test Results to Excel File...")
    # Convert data to a table
    df = pd.DataFrame(MKResults, columns=['Site Number', 'Trend', 'Hypothesis Rejected (T/F)', 'P-Value', 'Mean',
                                          'Median', 'Standard Deviation'])

    # Convert table to Excel file
    df.to_excel('Test-Results.xlsx', sheet_name='Sheet1')

    # Convert table to Excel file with specified formatting
    with pd.ExcelWriter('Test-Results.xlsx', engine='xlsxwriter') as writer:
        df.to_excel(writer, sheet_name='Sheet1', index=False)

        # Access the XlsxWriter workbook and worksheet objects
        workbook = writer.book
        worksheet = writer.sheets['Sheet1']

        # Create a number format for 10 decimal places
        number_format = workbook.add_format({'num_format': '0.0000000000'})

        # Apply the format to the 'Mean' column (5th column, index 4 in zero-based index)
        worksheet.set_column('E:E', None, number_format)

    # Create a number format for 10 decimal places
    number_format = workbook.add_format({'num_format': '0.0000000000'})

    # Apply the format to the 'Mean' column (5th column, index 4 in zero-based index)
    worksheet.set_column('E:E', None, number_format)

    print("Complete.")
    print("Formatting Data Types...")
    workbook = load_workbook("Test-Results.xlsx")
    worksheet = workbook["Sheet1"]
    row = 2

    # Format data to prep the Excel spreadsheet for Power BI Integration
    while row <= 1400:
        columnA = "A" + str(row)
        columnB = "B" + str(row)
        columnC = "C" + str(row)
        columnD = "D" + str(row)
        columnE = "E" + str(row)
        columnF = "F" + str(row)
        columnG = "G" + str(row)
        columnH = "H" + str(row)
        columnI = "I" + str(row)
        columnJ = "J" + str(row)

        worksheet[columnA].number_format = numbers.FORMAT_TEXT
        worksheet[columnB].number_format = numbers.FORMAT_TEXT
        worksheet[columnC].number_format = numbers.FORMAT_TEXT
        worksheet[columnD].number_format = numbers.FORMAT_TEXT
        worksheet[columnE].number_format = numbers.FORMAT_NUMBER
        worksheet[columnF].number_format = numbers.FORMAT_NUMBER
        worksheet[columnG].number_format = numbers.FORMAT_NUMBER
        worksheet[columnH].number_format = numbers.FORMAT_NUMBER
        worksheet[columnI].number_format = numbers.FORMAT_NUMBER
        worksheet[columnJ].number_format = numbers.FORMAT_NUMBER

        row += 1
    print("Complete.")
    workbook.save('Test-Results.xlsx')
    print("Workbook Saved.")

    formatResults()
    print("Sites not included: ", dontinclude)

    return 0


main()
