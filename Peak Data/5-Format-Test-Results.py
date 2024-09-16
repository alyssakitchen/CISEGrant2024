from openpyxl import load_workbook


def main():
    workbook = load_workbook("../all-sites.xlsx")
    worksheet = workbook["Sheet1"]

    site = []
    AllSites = []
    row = 2

    print("Pulling data for comparison...")
    # Pull Longitude, Latitude, and Site Name for every Site Number (1383 total)
    while row < 1384:
        site_no = "B"+str(row)
        station_nm = "C"+str(row)
        latitude = "D"+str(row)
        longitude = "E"+str(row)

        if worksheet[site_no] is not None:
            site.append(worksheet[site_no].value)
            site.append(worksheet[station_nm].value)
            site.append(worksheet[latitude].value)
            site.append(worksheet[longitude].value)

        AllSites.append(site)
        site = []
        row += 1

    print("Complete.")
    workbook2 = load_workbook("Test-Results.xlsx")
    worksheet2 = workbook2["Sheet1"]
    worksheet2.insert_cols(1)

    # Variable definitions
    row = 1
    columnA = "A" + str(row)
    columnI = "I" + str(row)
    columnJ = "J" + str(row)

    print("Adding titles for new data columns...")
    # Add titles for the new columns
    worksheet2[columnA].value = "Station Name"
    worksheet2[columnI].value = "Latitude"
    worksheet2[columnJ].value = "Longitude"

    print("Complete.")
    workbook2.save('Test-Results.xlsx')
    print("Workbook Saved.")

    # Find matching site numbers
    workingSites = []
    while row < 1400:
        row += 1
        columnB = "B" + str(row)
        # If we're not at the end of the data
        if worksheet2[columnB].value is not None:
            workingSites.append(worksheet2[columnB].value)
    #print(len(workingSites)) # debug

    print("Importing Latitude/Longitude data for dataset...")
    count = 0
    for index in AllSites:
        for i in workingSites:
            # If the site number at index i from working sites is equal to that from all the sites...
            if i == index[0]:
                columnA = "A" + str(count+2)
                columnI = "I" + str(count+2)
                columnJ = "J" + str(count+2)
                # ... add the lat/long data and site name to the correct column in Test-Results.xlsx
                worksheet2[columnA].value = index[1]
                worksheet2[columnI].value = index[2]
                worksheet2[columnJ].value = index[3]
                count += 1
    #print(count) # debug

    print("Complete.")
    workbook2.save('Test-Results.xlsx')
    print("Workbook Saved.")

    return 0


main()
