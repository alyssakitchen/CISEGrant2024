from openpyxl import load_workbook
import pymannkendall as mk
import statistics
from openpyxl.styles import numbers
import pandas as pd


def main():
    workbook = load_workbook("peak.xlsx")
    worksheet = workbook["Sheet1"]

    print("Importing Data...")

    # Initialize Lists and Variables
    siteNumbers = []
    Lists = []  # This is the list of lists
    data = []   # This is the temporary list that will be appended to Lists[] per each site number
    row = 3
    starting = "E" + str(row - 1)

    if not worksheet[starting].value:
        worksheet[starting].value = 0
    data.append(float(worksheet[starting].value))
    while row < 20000:
        # Initialize Changing Variables
        columnB = "B" + str(row)
        previous = "B" + str(row - 1)
        columnE = "E" + str(row)

        # Check to make sure we're not at the end of the data
        if worksheet[columnB] is not None:
            # If the site number hasn't changed
            if worksheet[columnB].value == worksheet[previous].value:
                if worksheet[columnE].value is None:
                    data.append(0.0)
                else:
                    data.append(float(worksheet[columnE].value))
            # If the site number has changed
            elif worksheet[columnB].value != worksheet[previous].value:
                if len(data) >= 80:     # Change to 80
                    Lists.append(data)
                    siteNumbers.append(worksheet[previous].value)
                    print("Data for site number", worksheet[previous].value, " added to dataset.")
                elif len(data) < 80:    # Change to 80
                    print("Not enough data for site number", worksheet[previous].value, "to include in report.")
                if worksheet[columnE].value is None:
                    data = [0.0]
                else:
                    value = str(worksheet[columnE].value)
                    data = [float(value)]
        row += 1

    print("Complete.")
    # print(len(siteNumbers)) #debug
    # print(len(Lists)) #debug
    index = 0

    MKResults = []
    individual = []

    print("Performing Tests...")
    # Import Results into a list (MKResults) of lists (individual)
    while index < len(siteNumbers):
        # Perform the Mann-Kendall test
        site = siteNumbers[index]
        result = mk.original_test(Lists[index], alpha=0.05)
        mean = statistics.mean(Lists[index])
        median = statistics.median(Lists[index])
        stdev = statistics.stdev(Lists[index])

        # Extract results
        trend = result.trend
        hypothesis = result.h
        p_value = result.p
        print("P:", p_value)

        # Add data to individual
        individual.append(site)
        individual.append(trend)
        individual.append(hypothesis)
        individual.append(p_value)
        individual.append(mean)
        individual.append(median)
        individual.append(stdev)

        # Append individual site results to MKResults
        MKResults.append(individual)
        index += 1
        individual = []
    print("Complete.")
    print(MKResults) # debug

    print("Importing Test Results to Excel File...")
    # Convert data to a table
    df = pd.DataFrame(MKResults, columns=['Site Number', 'Trend', 'Hypothesis Rejected (T/F)', 'P-Value', 'Mean',
                                          'Median', 'Standard Deviation'])

    # Convert table to Excel file
    df.to_excel('Test-Results.xlsx', sheet_name='Sheet1')

    # Convert table to Excel file with specified formatting
    with pd.ExcelWriter('Test-Results.xlsx', engine='xlsxwriter') as writer:
        df.to_excel(writer, sheet_name='Sheet1', index=False)

        # Access the XlsxWriter workbook and worksheet objects
        workbook = writer.book
        worksheet = writer.sheets['Sheet1']

        # Create a number format for 10 decimal places
        number_format = workbook.add_format({'num_format': '0.0000000000'})

        # Apply the format to the 'Mean' column (5th column, index 4 in zero-based index)
        worksheet.set_column('E:E', None, number_format)

    # Create a number format for 10 decimal places
    number_format = workbook.add_format({'num_format': '0.0000000000'})

    # Apply the format to the 'Mean' column (5th column, index 4 in zero-based index)
    worksheet.set_column('E:E', None, number_format)

    print("Complete.")
    print("Formatting Data Types...")
    workbook = load_workbook("Test-Results.xlsx")
    worksheet = workbook["Sheet1"]
    row = 2

    # Format data to prep the Excel spreadsheet for Power BI Integration
    while row <= 1400:
        columnA = "A" + str(row)
        columnB = "B" + str(row)
        columnC = "C" + str(row)
        columnD = "D" + str(row)
        columnE = "E" + str(row)
        columnF = "F" + str(row)
        columnG = "G" + str(row)
        columnH = "H" + str(row)
        columnI = "I" + str(row)
        columnJ = "J" + str(row)

        worksheet[columnA].number_format = numbers.FORMAT_TEXT
        worksheet[columnB].number_format = numbers.FORMAT_TEXT
        worksheet[columnC].number_format = numbers.FORMAT_TEXT
        worksheet[columnD].number_format = numbers.FORMAT_TEXT
        worksheet[columnE].number_format = numbers.FORMAT_NUMBER
        worksheet[columnF].number_format = numbers.FORMAT_NUMBER
        worksheet[columnG].number_format = numbers.FORMAT_NUMBER
        worksheet[columnH].number_format = numbers.FORMAT_NUMBER
        worksheet[columnI].number_format = numbers.FORMAT_NUMBER
        worksheet[columnJ].number_format = numbers.FORMAT_NUMBER

        row += 1
    print("Complete.")
    workbook.save('Test-Results.xlsx')
    print("Workbook Saved.")

    return 0


main()
