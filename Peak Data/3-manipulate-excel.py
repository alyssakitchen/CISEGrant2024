from openpyxl import load_workbook
from openpyxl.styles import numbers


def main():
    workbook = load_workbook("peak.xlsx")
    worksheet = workbook["Sheet1"]
    row = 2

    print("Reformatting Data Types...")
    # Convert Columns A, B, C, D, E, & F to Text Data Type and G & H to Number Data Type
    while row <= 20000:
        columnA = "A" + str(row)
        columnB = "B" + str(row)
        columnC = "C" + str(row)
        columnD = "D" + str(row)
        columnE = "E" + str(row)
        columnF = "F" + str(row)
        columnG = "G" + str(row)
        columnH = "H" + str(row)

        worksheet[columnA].number_format = numbers.FORMAT_TEXT
        worksheet[columnB].number_format = numbers.FORMAT_TEXT
        worksheet[columnC].number_format = numbers.FORMAT_TEXT
        worksheet[columnD].number_format = numbers.FORMAT_TEXT
        worksheet[columnE].number_format = numbers.FORMAT_NUMBER
        worksheet[columnF].number_format = numbers.FORMAT_NUMBER
        worksheet[columnG].number_format = numbers.FORMAT_NUMBER
        worksheet[columnH].number_format = numbers.FORMAT_NUMBER

        row += 1
    print("Complete.")

    workbook.save('peak.xlsx')
    print("Workbook Saved.")

    print("Finding where peak_cd contains 5 or 6...")
    # Remove all rows where peak_cd contains 5 or 6
    rows_to_delete = []
    for row in worksheet.iter_rows(min_row=2, min_col=6, max_col=6):  # assuming data starts from row 2 in column H
        cell_value = row[0].value
        if cell_value is not None and ('5' in str(cell_value) or '6' in str(cell_value)):
            rows_to_delete.append(row[0].row)

    print("Removing rows (this may take several minutes)...")
    for index in reversed(rows_to_delete):
        worksheet.delete_rows(index)
    print("Complete.")

    # Save the workbook
    workbook.save('peak.xlsx')
    print("Workbook Saved.")

    return 0


main()
